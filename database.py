import sqlite3
import bcrypt
import numpy as np
from datetime import datetime, date as date_class

DB_FILE = "attendance.db"

# PASSWORD HASHING (bcrypt)

def hash_password(plain_password: str) -> str:
    """Hash a plain-text password using bcrypt. Returns a UTF-8 string for storage."""
    # Cost 10 keeps password hashing secure for this local school project
    # while allowing Excel imports to finish promptly on typical laptops.
    # bcrypt can still verify hashes made with any previous cost factor.
    salt = bcrypt.gensalt(rounds=10)
    hashed = bcrypt.hashpw(plain_password.encode("utf-8"), salt)
    return hashed.decode("utf-8")


def check_password(plain_password: str, hashed_password: str) -> bool:
    """Verify a plain-text password against a stored bcrypt hash."""
    return bcrypt.checkpw(
        plain_password.encode("utf-8"),
        hashed_password.encode("utf-8")
    )


#  DATABASE CREATION 

def create_database():
    """Create all tables for the Smart Attendance System"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    # STUDENTS TABLE
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS students (
            student_id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            course TEXT NOT NULL,
            year INTEGER NOT NULL,
            email TEXT,
            password TEXT
        )
    """)

    #Student faces
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS student_faces (
            face_id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id TEXT NOT NULL,
            image_path TEXT,
            embedding BLOB NOT NULL,
            FOREIGN KEY(student_id) REFERENCES students(student_id)
        )
    """)

    # UNRECOGNIZED LOGS TABLE (Saves snapshots of intruders or unregistered users)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS unrecognized_logs (
            log_id INTEGER PRIMARY KEY AUTOINCREMENT,
            image_path TEXT NOT NULL,
            timestamp TEXT NOT NULL
        )
    """)

    # SUBJECTS TABLE
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS subjects (
            subject_id TEXT PRIMARY KEY,
            subject_name TEXT NOT NULL,
            course TEXT NOT NULL,
            year INTEGER NOT NULL,
            room TEXT NOT NULL,
            teacher_name TEXT NOT NULL
        )
    """)

    # ATTENDANCE TABLE
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id TEXT NOT NULL,
            subject_id TEXT NOT NULL,
            date TEXT NOT NULL,
            time TEXT NOT NULL,
            UNIQUE(student_id, subject_id, date),
            FOREIGN KEY(student_id) REFERENCES students(student_id),
            FOREIGN KEY(subject_id) REFERENCES subjects(subject_id)
        )
    """)

    # SESSIONS TABLE
    # routine_id links a session back to its slot on the weekly timetable.
    # routine_id = NULL means an "alternate" / extra class not on the routine.
    # SQLite treats every NULL as distinct for UNIQUE purposes, so this
    # constraint still blocks duplicate *routine* sessions per day while
    # allowing any number of alternate-class sessions on the same day.
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS sessions (
            session_id INTEGER PRIMARY KEY AUTOINCREMENT,
            subject_id TEXT NOT NULL,
            date TEXT NOT NULL,
            room TEXT NOT NULL,
            start_time TEXT,
            end_time TEXT,
            notes TEXT,
            routine_id INTEGER,
            is_alternate INTEGER NOT NULL DEFAULT 0,
            UNIQUE(subject_id, date, routine_id),
            FOREIGN KEY(subject_id) REFERENCES subjects(subject_id),
            FOREIGN KEY(routine_id) REFERENCES routine(routine_id)
        )
    """)

    # ACTIVE_SESSIONS TABLE
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS active_sessions (
            session_id INTEGER PRIMARY KEY,
            opened_at TEXT NOT NULL,
            FOREIGN KEY(session_id) REFERENCES sessions(session_id)
        )
    """)

    # TEACHERS TABLE — passwords stored as bcrypt hashes
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS teachers (
            teacher_id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            email TEXT,
            phone TEXT,
            department TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)

    _create_communication_tables(cursor)

    # ROUTINE TABLE — class timetable (drives the routine view shown to students/teachers)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS routine (
            routine_id INTEGER PRIMARY KEY AUTOINCREMENT,
            course TEXT NOT NULL,
            year INTEGER NOT NULL,
            section TEXT NOT NULL,
            day TEXT NOT NULL,
            start_time TEXT NOT NULL,
            end_time TEXT NOT NULL,
            subject_id TEXT NOT NULL,
            room TEXT,
            FOREIGN KEY(subject_id) REFERENCES subjects(subject_id)
        )
    """)

    conn.commit()
    conn.close()
    print(f"Database created: {DB_FILE}")
    print("  - students table")
    print("  - subjects table")
    print("  - attendance table")
    print("  - sessions table")
    print("  - active_sessions table")
    print("  - teachers table (bcrypt passwords)")
    print("  - routine table")
    print("  - chat_messages table")
    print("  - notifications table")


def _create_communication_tables(cursor):
    """Create the schema used by the notification and two-way chat pages."""
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS notifications (
            notification_id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id TEXT NOT NULL,
            subject_id TEXT NOT NULL,
            attendance_date TEXT NOT NULL,
            message TEXT NOT NULL,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            expires_at TEXT NOT NULL,
            read_at TEXT,
            UNIQUE(student_id, subject_id, attendance_date),
            FOREIGN KEY(student_id) REFERENCES students(student_id),
            FOREIGN KEY(subject_id) REFERENCES subjects(subject_id)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS chat_messages (
            message_id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id TEXT NOT NULL,
            teacher_id INTEGER NOT NULL,
            subject_id TEXT NOT NULL,
            sender_role TEXT NOT NULL CHECK(sender_role IN ('student', 'teacher')),
            body TEXT NOT NULL,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            read_at TEXT,
            unsent_at TEXT,
            FOREIGN KEY(student_id) REFERENCES students(student_id),
            FOREIGN KEY(teacher_id) REFERENCES teachers(teacher_id),
            FOREIGN KEY(subject_id) REFERENCES subjects(subject_id)
        )
    """)
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_notifications_student ON notifications(student_id, expires_at, read_at)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_chat_student ON chat_messages(student_id, created_at)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_chat_teacher ON chat_messages(teacher_id, subject_id, created_at)")


def migrate_communication_tables():
    """Upgrade the earlier prototype tables without discarding its messages."""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    try:
        tables = {row[0] for row in cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")}
        migrate_notifications = False
        migrate_chat = False
        if "notifications" in tables:
            notification_columns = {row[1] for row in cursor.execute("PRAGMA table_info(notifications)")}
            if "attendance_date" not in notification_columns:
                cursor.execute("ALTER TABLE notifications RENAME TO notifications_legacy")
                migrate_notifications = True
        if "chat_messages" in tables:
            chat_columns = {row[1] for row in cursor.execute("PRAGMA table_info(chat_messages)")}
            if "student_id" not in chat_columns:
                cursor.execute("ALTER TABLE chat_messages RENAME TO chat_messages_legacy")
                migrate_chat = True
        _create_communication_tables(cursor)
        if migrate_notifications:
            cursor.execute("""
                INSERT INTO notifications
                (notification_id, student_id, subject_id, attendance_date, message, created_at, expires_at, read_at)
                SELECT notification_id, student_id, COALESCE(subject_id, ''), date(created_at), message,
                       created_at, datetime(created_at, '+1 day'),
                       CASE WHEN is_read THEN created_at ELSE NULL END
                FROM notifications_legacy
            """)
        if migrate_chat:
            cursor.execute("""
                INSERT INTO chat_messages
                (message_id, student_id, teacher_id, subject_id, sender_role, body, created_at, read_at)
                SELECT message_id,
                       CASE WHEN sender_type='student' THEN sender_id ELSE receiver_id END,
                       CAST(CASE WHEN sender_type='teacher' THEN sender_id ELSE receiver_id END AS INTEGER),
                       subject_id, sender_type, message, sent_at,
                       CASE WHEN is_read THEN sent_at ELSE NULL END
                FROM chat_messages_legacy
                WHERE sender_type IN ('student', 'teacher')
            """)
        conn.commit()
    finally:
        conn.close()


def create_attendance_notification(student_id, subject_id, attendance_date=None):
    """Create one unread attendance alert which expires after 24 hours."""
    attendance_date = attendance_date or str(date_class.today())
    conn = sqlite3.connect(DB_FILE)
    try:
        subject = conn.execute("SELECT subject_name FROM subjects WHERE subject_id=?", (subject_id,)).fetchone()
        subject_name = subject[0] if subject else subject_id
        conn.execute(
            """INSERT OR IGNORE INTO notifications
               (student_id, subject_id, attendance_date, message, expires_at)
               VALUES (?, ?, ?, ?, datetime('now', '+1 day'))""",
            (student_id, subject_id, attendance_date,
             f"Your attendance has been marked for {subject_id} — {subject_name}.")
        )
        conn.commit()
    finally:
        conn.close()





def save_attendance_record(student_id: str, subject_id: str):
    """Inserts a safe transaction stamp into the attendance engine log."""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    today_str = str(date_class.today())
    time_str = datetime.now().strftime("%H:%M:%S")
    
    try:
        cursor.execute("""
            INSERT INTO attendance (student_id, subject_id, date, time)
            VALUES (?, ?, ?, ?)
        """, (student_id, subject_id, today_str, time_str))
        conn.commit()
        success = True
    except sqlite3.IntegrityError:
        success = False # Handled by database unique constraint check
    finally:
        conn.close()
    if success:
        create_attendance_notification(student_id, subject_id, today_str)
    return success

def log_unrecognized_detection(image_path: str):
    """Records an unmapped entity capture trail into audit tracking."""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    now_str = datetime.now().isoformat()
    cursor.execute("""
        INSERT INTO unrecognized_logs (image_path, timestamp)
        VALUES (?, ?)
    """, (image_path, now_str))
    conn.commit()
    conn.close()

def get_current_active_subject():
    """Determines which class is currently running based on active sessions."""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT s.subject_id FROM active_sessions axs
        JOIN sessions s ON axs.session_id = s.session_id
        LIMIT 1
    """)
    row = cursor.fetchone()
    conn.close()
    return row[0] if row else "DEFAULT_SUB"

# EXCEL IMPORT 

def import_students_from_excel(filepath: str) -> dict:
    """
    Import students from an Excel file.

    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("openpyxl not installed. Run: pip install openpyxl")
        return {"inserted": 0, "skipped": 0, "errors": ["openpyxl not installed"]}

    wb = load_workbook(filepath)
    ws = wb.active

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    inserted = 0
    skipped = 0
    errors = []

    # Skip the header row
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    for i, row in enumerate(rows, start=2):
        # Skip completely empty rows
        if not any(row):
            continue

        # Validate column count
        if len(row) < 4:
            errors.append(f"Row {i}: not enough columns (expected 4, got {len(row)})")
            continue

        student_id, name, course, year = row[0], row[1], row[2], row[3]
        email    = row[4] if len(row) > 4 else None
        password = row[5] if len(row) > 5 else None

        # Validate required fields
        if not student_id or not name or not course or not year:
            errors.append(f"Row {i}: missing required field — {row}")
            continue

        try:
            year = int(year)
        except (ValueError, TypeError):
            errors.append(f"Row {i}: 'year' must be a number, got '{year}'")
            continue

        try:
            hashed_pw = hash_password(str(password)) if password else None
            cursor.execute(
                "INSERT OR IGNORE INTO students (student_id, name, course, year, email, password) VALUES (?, ?, ?, ?, ?, ?)",
                (str(student_id).strip(), str(name).strip(), str(course).strip(), year,
                 str(email).strip() if email else None, hashed_pw)
            )
            if cursor.rowcount > 0:
                inserted += 1
            else:
                skipped += 1  # Student ID already exists
        except sqlite3.Error as e:
            errors.append(f"Row {i}: database error — {e}")

    conn.commit()
    conn.close()

    print(f"Students import complete — inserted: {inserted}, skipped: {skipped}, errors: {len(errors)}")
    for err in errors:
        print(f" {err}")

    return {"inserted": inserted, "skipped": skipped, "errors": errors}


def import_teachers_from_excel(filepath: str) -> dict:
    """
    Import teachers from an Excel file.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print(" openpyxl not installed. Run: pip install openpyxl")
        return {"inserted": 0, "skipped": 0, "errors": ["openpyxl not installed"]}

    wb = load_workbook(filepath)
    ws = wb.active

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    inserted = 0
    skipped = 0
    errors = []

    rows = list(ws.iter_rows(min_row=2, values_only=True))

    for i, row in enumerate(rows, start=2):
        if not any(row):
            continue

        if len(row) < 6:
            errors.append(f"Row {i}: not enough columns (expected 6, got {len(row)})")
            continue

        name, email, phone, department, username, password = (
            row[0], row[1], row[2], row[3], row[4], row[5]
        )

        if not name or not username or not password:
            errors.append(f"Row {i}: name, username, and password are required")
            continue

        # Hash the plain-text password from the sheet
        hashed = hash_password(str(password))

        try:
            cursor.execute(
                """INSERT INTO teachers (name, username, password, email, phone, department)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(username) DO UPDATE SET
                    name=excluded.name,
                    email=excluded.email,
                    phone=excluded.phone,
                    department=excluded.department""",
                (
                    str(name).strip(),
                    str(username).strip(),
                    hashed,
                    str(email).strip() if email else None,
                    str(phone).strip() if phone else None,
                    str(department).strip() if department else None,
                )
            )
            if cursor.rowcount > 0:
                inserted += 1
            else:
                skipped += 1  # Username already exists
        except sqlite3.Error as e:
            errors.append(f"Row {i}: database error — {e}")

    conn.commit()
    conn.close()

    print(f"Teachers import complete — inserted: {inserted}, skipped: {skipped}, errors: {len(errors)}")
    for err in errors:
        print(f"  ⚠ {err}")

    return {"inserted": inserted, "skipped": skipped, "errors": errors}


#  STUDENT QUERY FUNCTIONS 

def get_students_for_course(course, year):
    """Get all students in a course/year combination"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute(
        "SELECT student_id, name FROM students WHERE course = ? AND year = ?",
        (course, year)
    )
    result = cursor.fetchall()
    conn.close()
    return result


def get_subjects_for_student(student_id):
    """Get all subjects for a student's course/year"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT course, year FROM students WHERE student_id = ?", (student_id,))
    student = cursor.fetchone()
    if not student:
        conn.close()
        return []
    course, year = student
    cursor.execute(
        "SELECT subject_id, subject_name, room, teacher_name FROM subjects WHERE course = ? AND year = ?",
        (course, year)
    )
    result = cursor.fetchall()
    conn.close()
    return result


def get_all_subjects():
    """Get all subjects (for teacher view)"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT subject_id, subject_name, room, teacher_name FROM subjects")
    result = cursor.fetchall()
    conn.close()
    return result


def get_attendance_for_student_subject(student_id, subject_id):
    """Get all attendance records for a student in a specific subject"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute(
        "SELECT date, time FROM attendance WHERE student_id = ? AND subject_id = ? ORDER BY date DESC",
        (student_id, subject_id)
    )
    result = cursor.fetchall()
    conn.close()
    return result


def get_attendance_count_for_student_subject(student_id, subject_id):
    """Get count of days attended for a student in a subject"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute(
        "SELECT COUNT(DISTINCT date) FROM attendance WHERE student_id = ? AND subject_id = ?",
        (student_id, subject_id)
    )
    result = cursor.fetchone()[0]
    conn.close()
    return result


def get_total_class_days(subject_id):
    """Get number of unique class days for a subject"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute(
        "SELECT COUNT(DISTINCT date) FROM attendance WHERE subject_id = ?",
        (subject_id,)
    )
    result = cursor.fetchone()[0]
    conn.close()
    return result if result > 0 else 1


def get_todays_attendance(subject_id):
    """Get all students who have scanned in today for a subject"""
    today = str(date_class.today())
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute(
        """SELECT DISTINCT s.student_id, s.name, a.time
           FROM attendance a
           JOIN students s ON a.student_id = s.student_id
           WHERE a.subject_id = ? AND a.date = ?
           ORDER BY a.time ASC""",
        (subject_id, today)
    )
    result = cursor.fetchall()
    conn.close()
    return result


def get_full_attendance_report(subject_id):
    """Get full attendance summary for all students in a subject"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute(
        "SELECT DISTINCT course, year FROM subjects WHERE subject_id = ?",
        (subject_id,)
    )
    course_year = cursor.fetchone()
    if not course_year:
        conn.close()
        return []
    course, year = course_year
    cursor.execute(
        """SELECT s.student_id, s.name, COUNT(DISTINCT a.date) as days_present
           FROM students s
           LEFT JOIN attendance a ON s.student_id = a.student_id AND a.subject_id = ?
           WHERE s.course = ? AND s.year = ?
           GROUP BY s.student_id
           ORDER BY s.name ASC""",
        (subject_id, course, year)
    )
    result = cursor.fetchall()
    conn.close()
    return result


def student_exists(student_id):
    """Check if a student exists in the database"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT 1 FROM students WHERE student_id = ?", (student_id,))
    result = cursor.fetchone() is not None
    conn.close()
    return result


def get_student_name(student_id):
    """Get student name by ID"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM students WHERE student_id = ?", (student_id,))
    result = cursor.fetchone()
    conn.close()
    return result[0] if result else None


#  TEACHER AUTH FUNCTIONS 

def verify_teacher_login(username, password):
    """
    Verify teacher login credentials using bcrypt.
    Returns teacher dict on success, None on failure.
    """
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute(
        "SELECT teacher_id, name, department, email, password FROM teachers WHERE username = ?",
        (username,)
    )
    result = cursor.fetchone()
    conn.close()

    if result and check_password(password, result[4]):
        return {
            "teacher_id": result[0],
            "name": result[1],
            "department": result[2],
            "email": result[3],
        }
    return None


def get_teacher_by_id(teacher_id):
    """Get teacher details by ID"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute(
        "SELECT teacher_id, name, username, email, phone, department FROM teachers WHERE teacher_id = ?",
        (teacher_id,)
    )
    result = cursor.fetchone()
    conn.close()
    if result:
        return {
            "teacher_id": result[0],
            "name": result[1],
            "username": result[2],
            "email": result[3],
            "phone": result[4],
            "department": result[5],
        }
    return None


def get_all_teachers():
    """Get all teachers"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT teacher_id, name, username, email, phone, department FROM teachers")
    result = cursor.fetchall()
    conn.close()
    return result


#  SESSION MANAGEMENT FUNCTIONS 

def create_or_update_session(subject_id, date, room, start_time=None, end_time=None, notes="", routine_id=None):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    try:
        cursor.execute(
            """INSERT INTO sessions (subject_id, date, room, start_time, end_time, notes, routine_id)
               VALUES (?, ?, ?, ?, ?, ?, ?)
               ON CONFLICT(subject_id, date, routine_id) DO UPDATE SET
                   room=excluded.room, start_time=excluded.start_time,
                   end_time=excluded.end_time, notes=excluded.notes""",
            (subject_id, date, room, start_time, end_time, notes, routine_id)
        )
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"Error creating session: {e}")
        conn.close()
        return False


def get_session_for_date(subject_id, date):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute(
        """SELECT session_id, subject_id, date, room, start_time, end_time, notes
           FROM sessions WHERE subject_id = ? AND date = ?""",
        (subject_id, date)
    )
    result = cursor.fetchone()
    conn.close()
    return result


def get_room_for_subject_date(subject_id, date):
    session = get_session_for_date(subject_id, date)
    if session:
        return session[3]
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT room FROM subjects WHERE subject_id = ?", (subject_id,))
    result = cursor.fetchone()
    conn.close()
    return result[0] if result else None


def open_session(subject_id, date_str=None):
    if date_str is None:
        date_str = str(date_class.today())
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    session = get_session_for_date(subject_id, date_str)
    if not session:
        default_room = conn.execute(
            "SELECT room FROM subjects WHERE subject_id = ?", (subject_id,)
        ).fetchone()
        if not default_room:
            conn.close()
            return False
        cursor.execute(
            "INSERT INTO sessions (subject_id, date, room) VALUES (?, ?, ?)",
            (subject_id, date_str, default_room[0])
        )
        conn.commit()
        session = get_session_for_date(subject_id, date_str)
    if not session:
        conn.close()
        return False
    session_id = session[0]
    now = datetime.now().isoformat()
    try:
        cursor.execute(
            "INSERT INTO active_sessions (session_id, opened_at) VALUES (?, ?)",
            (session_id, now)
        )
        conn.commit()
        conn.close()
        return True
    except sqlite3.IntegrityError:
        conn.close()
        return False


def close_session(subject_id, date_str=None):
    if date_str is None:
        date_str = str(date_class.today())
    session = get_session_for_date(subject_id, date_str)
    if not session:
        return False
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM active_sessions WHERE session_id = ?", (session[0],))
    conn.commit()
    conn.close()
    return True


def is_session_open(subject_id, date_str=None):
    if date_str is None:
        date_str = str(date_class.today())
    session = get_session_for_date(subject_id, date_str)
    if not session:
        return False
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT 1 FROM active_sessions WHERE session_id = ?", (session[0],))
    result = cursor.fetchone() is not None
    conn.close()
    return result


def get_all_sessions_for_subject(subject_id, limit=30):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute(
        """SELECT session_id, subject_id, date, room, start_time, end_time, notes
           FROM sessions WHERE subject_id = ?
           ORDER BY date DESC LIMIT ?""",
        (subject_id, limit)
    )
    result = cursor.fetchall()
    conn.close()
    return result


def migrate_sessions_table():
    """
    One-time migration for existing databases created before routine_id /
    is_alternate / alternate-class support was added. Safe to call every
    startup — it no-ops once the new columns exist.
    """
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("PRAGMA table_info(sessions)")
    cols = {row[1] for row in cursor.fetchall()}
    if "routine_id" in cols and "is_alternate" in cols:
        conn.close()
        return  # already migrated

    print("Migrating sessions table to support routine_id / alternate classes...")
    cursor.execute("ALTER TABLE sessions RENAME TO sessions_old")
    cursor.execute("""
        CREATE TABLE sessions (
            session_id INTEGER PRIMARY KEY AUTOINCREMENT,
            subject_id TEXT NOT NULL,
            date TEXT NOT NULL,
            room TEXT NOT NULL,
            start_time TEXT,
            end_time TEXT,
            notes TEXT,
            routine_id INTEGER,
            is_alternate INTEGER NOT NULL DEFAULT 0,
            UNIQUE(subject_id, date, routine_id),
            FOREIGN KEY(subject_id) REFERENCES subjects(subject_id),
            FOREIGN KEY(routine_id) REFERENCES routine(routine_id)
        )
    """)
    cursor.execute("""
        INSERT INTO sessions (session_id, subject_id, date, room, start_time, end_time, notes, routine_id, is_alternate)
        SELECT session_id, subject_id, date, room, start_time, end_time, notes, NULL, 0
        FROM sessions_old
    """)
    cursor.execute("DROP TABLE sessions_old")
    # active_sessions references session_id, which is preserved, so it needs no changes
    conn.commit()
    conn.close()
    print("Migration complete.")


def get_routine_slot_for_today(subject_id, course, year, day, section="CE-II/II"):
    """
    Find the routine row (if any) that matches this subject on this day —
    used to tie a normal "open session" to its timetable slot.
    Returns the routine_id, or None if this subject isn't scheduled today
    (in which case opening a session for it should be treated as alternate).
    """
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute(
        """SELECT routine_id FROM routine
           WHERE subject_id=? AND course=? AND year=? AND day=? AND section=?
           LIMIT 1""",
        (subject_id, course, year, day, section)
    )
    row = cursor.fetchone()
    conn.close()
    return row[0] if row else None


def open_routine_or_alternate_session(subject_id, room=None, force_alternate=False):
    """
    Opens a session for `subject_id` today.
    - If the subject is on today's routine and force_alternate is False,
      the session is linked to that routine slot (normal class).
    - Otherwise (force_alternate=True, or subject isn't on today's routine),
      a brand-new alternate session row is created — multiple alternate
      sessions for the same subject/day are allowed.
    Returns the session_id.
    """
    today = date_class.today()
    today_str = str(today)
    day_name = today.strftime("%A")

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    sub = cursor.execute(
        "SELECT course, year, room FROM subjects WHERE subject_id=?", (subject_id,)
    ).fetchone()
    if not sub:
        conn.close()
        return None
    course, year, default_room = sub
    room = room or default_room

    routine_id = None
    if not force_alternate:
        routine_id = get_routine_slot_for_today(subject_id, course, year, day_name)

    is_alternate = 1 if routine_id is None else 0
    now_iso = datetime.now().isoformat()

    if routine_id is not None:
        # Normal routine session — reuse if it already exists today
        cursor.execute(
            "SELECT session_id FROM sessions WHERE subject_id=? AND date=? AND routine_id=?",
            (subject_id, today_str, routine_id)
        )
        existing = cursor.fetchone()
        if existing:
            session_id = existing[0]
        else:
            cursor.execute(
                """INSERT INTO sessions (subject_id, date, room, start_time, routine_id, is_alternate)
                   VALUES (?, ?, ?, ?, ?, 0)""",
                (subject_id, today_str, room, now_iso, routine_id)
            )
            conn.commit()
            session_id = cursor.lastrowid
    else:
        # Alternate class — always a new row, since routine_id is NULL
        # and we want a fresh session even if one was opened/closed earlier today
        cursor.execute(
            """INSERT INTO sessions (subject_id, date, room, start_time, routine_id, is_alternate)
               VALUES (?, ?, ?, ?, NULL, 1)""",
            (subject_id, today_str, room, now_iso)
        )
        conn.commit()
        session_id = cursor.lastrowid

    try:
        cursor.execute(
            "INSERT INTO active_sessions (session_id, opened_at) VALUES (?, ?)",
            (session_id, now_iso)
        )
        conn.commit()
    except sqlite3.IntegrityError:
        pass  # already active

    conn.close()
    return session_id

def send_chat_message(sender_id,
                      sender_type,
                      receiver_id,
                      receiver_type,
                      subject_id,
                      message):

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO chat_messages
        (sender_id,
         sender_type,
         receiver_id,
         receiver_type,
         subject_id,
         message)

        VALUES (?, ?, ?, ?, ?, ?)
    """, (
        sender_id,
        sender_type,
        receiver_id,
        receiver_type,
        subject_id,
        message
    ))

    conn.commit()
    conn.close()

def get_chat(student_id, teacher_id, subject_id):

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    cursor.execute("""

        SELECT
            sender_id,
            sender_type,
            message,
            sent_at

        FROM chat_messages

        WHERE

        subject_id=?

        AND

        (
            (sender_id=? AND receiver_id=?)

            OR

            (sender_id=? AND receiver_id=?)
        )

        ORDER BY sent_at ASC

    """,
    (
        subject_id,
        student_id,
        teacher_id,
        teacher_id,
        student_id
    ))

    rows = cursor.fetchall()

    conn.close()

    return rows

def get_messages_by_subject(subject_id):

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    cursor.execute("""

        SELECT *

        FROM chat_messages

        WHERE subject_id=?

        ORDER BY sent_at DESC

    """,(subject_id,))

    rows = cursor.fetchall()

    conn.close()

    return rows

def get_teacher_messages(teacher_id):

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    cursor.execute("""

        SELECT *

        FROM chat_messages

        WHERE receiver_id=?
           OR sender_id=?

        ORDER BY sent_at DESC

    """,(teacher_id,teacher_id))

    rows = cursor.fetchall()

    conn.close()

    return rows


def add_notification(student_id,
                     subject_id,
                     title,
                     message):

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    cursor.execute("""

        INSERT INTO notifications

        (
            student_id,
            subject_id,
            title,
            message
        )

        VALUES

        (?, ?, ?, ?)

    """,(student_id,
         subject_id,
         title,
         message))

    conn.commit()
    conn.close()


def get_notifications(student_id):

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    cursor.execute("""

        SELECT

        notification_id,
        title,
        message,
        created_at,
        is_read

        FROM notifications

        WHERE student_id=?

        ORDER BY created_at DESC

    """,(student_id,))

    rows = cursor.fetchall()

    conn.close()

    return rows

def get_unread_notification_count(student_id):

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    cursor.execute("""

        SELECT COUNT(*)

        FROM notifications

        WHERE student_id=?

        AND is_read=0

    """,(student_id,))

    count = cursor.fetchone()[0]

    conn.close()

    return count

def get_unread_chat_count(user_id):

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    cursor.execute("""

        SELECT COUNT(*)

        FROM chat_messages

        WHERE receiver_id=?

        AND is_read=0

    """,(user_id,))

    count = cursor.fetchone()[0]

    conn.close()

    return count



def close_active_session_by_id(session_id):
    """Closes one specific session (used since alternate classes can have
    multiple open session_ids for the same subject/day)."""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE sessions SET end_time=? WHERE session_id=?",
        (datetime.now().isoformat(), session_id)
    )
    cursor.execute("DELETE FROM active_sessions WHERE session_id=?", (session_id,))
    conn.commit()
    conn.close()


def get_open_sessions_for_subject_today(subject_id):
    """All currently-open session_ids for a subject today (normally just one,
    but can be more than one if an alternate class is also running)."""
    today_str = str(date_class.today())
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute(
        """SELECT s.session_id, s.routine_id, s.is_alternate, s.start_time
           FROM sessions s
           JOIN active_sessions a ON a.session_id = s.session_id
           WHERE s.subject_id=? AND s.date=?""",
        (subject_id, today_str)
    )
    result = cursor.fetchall()
    conn.close()
    return result


#  ROUTINE FUNCTIONS 

def get_routine_for_course(course, year, section="CE-II/II"):
    """Get the weekly timetable for a course/year/section, ordered by day and time."""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute(
        """SELECT r.day, r.start_time, r.end_time, r.subject_id,
                  s.subject_name, r.room, s.teacher_name
           FROM routine r
           JOIN subjects s ON r.subject_id = s.subject_id
           WHERE r.course = ? AND r.year = ? AND r.section = ?
           ORDER BY
               CASE r.day
                   WHEN 'Sunday' THEN 0 WHEN 'Monday' THEN 1 WHEN 'Tuesday' THEN 2
                   WHEN 'Wednesday' THEN 3 WHEN 'Thursday' THEN 4 WHEN 'Friday' THEN 5
                   WHEN 'Saturday' THEN 6 ELSE 7 END,
               r.start_time""",
        (course, year, section)
    )
    result = cursor.fetchall()
    conn.close()
    return result


#  SEED DATA — CE-II/II (KU DoCSE routine, effective May 14, 2026) 

def populate_ce2_data():
    """
    Seed the subjects and routine tables for CE-II/II based on the official
    KU DoCSE routine (Group: II CE-II/II).
    Safe to re-run — uses INSERT OR IGNORE / clears old CE-II/II routine rows first.
    """
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    course, year, section = "CE", 2, "CE-II/II"

    subjects = [
        ("COMP232", "Database Management Systems",      course, year, "9-406", "Mr. Rajan Thapa"),
        ("COMP204", "Data Communication and Networking", course, year, "10-108", "Mr. Josh Karki"),
        ("COMP231", "Microprocessor and Assembly Language", course, year, "9-405", "Prof. Dr. Krishna Basnet"),
        ("MATH207", "Differential Equations and Complex Variables",           course, year, "9-305", "Mohan Bahadur Rai"),
        ("MCSC202", "Numerical Methods",    course, year, "9-405", "Dr. Kishor Poudel"),
    ]
    for sub in subjects:
        cursor.execute(
                """INSERT INTO subjects (subject_id, subject_name, course, year, room, teacher_name)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(subject_id) DO UPDATE SET
                    subject_name=excluded.subject_name,
                    course=excluded.course,
                    year=excluded.year,
                    room=excluded.room,
                    teacher_name=excluded.teacher_name""",
                sub
            )

    cursor.execute("DELETE FROM routine WHERE course = ? AND year = ? AND section = ?", (course, year, section))

    # (day, start, end, subject_id, room)
    routine_rows = [
        ("Tuesday",   "09:00", "11:00", "COMP232", "9-406"),
        ("Tuesday",   "12:00", "14:00", "COMP204", "10-108"),
        ("Wednesday", "11:00", "13:00", "MATH207", "9-305"),
        ("Wednesday", "14:00", "16:00", "COMP204", "9-405"),
        ("Thursday",  "09:00", "11:00", "MCSC202", "9-405"),
        ("Thursday",  "12:00", "14:00", "COMP231", "9-405"),
        ("Thursday",  "14:00", "16:00", "MATH207", "9-405"),
        ("Friday",    "09:00", "11:00", "COMP232", "9-405"),
        ("Friday",    "12:00", "14:00", "COMP231", "9-312"),
        ("Friday",    "14:00", "16:00", "MCSC202", "9-406"),
    ]
    for day, start, end, subject_id, room in routine_rows:
        cursor.execute(
            """INSERT INTO routine (course, year, section, day, start_time, end_time, subject_id, room)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (course, year, section, day, start, end, subject_id, room)
        )

    conn.commit()
    conn.close()
    print("CE-II/II subjects and routine seeded.")


if __name__ == "__main__":
    print("Initializing Smart Attendance System database...")
    create_database()
    print("Database ready.")
    print()
    print("To import data from Excel, run: python setup.py")