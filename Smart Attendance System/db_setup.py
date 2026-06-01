import sqlite3
import bcrypt
from datetime import datetime, date as date_class


# SQLite database file
DB_FILE = "attendance.db"


# PASSWORD HASHING (bcrypt)


def hash_password(plain_password: str) -> str:
    """Hash a plain-text password using bcrypt. Returns a UTF-8 string for storage."""
    salt = bcrypt.gensalt()
    hashed = bcrypt.hashpw(plain_password.encode("utf-8"), salt)
    return hashed.decode("utf-8")




def check_password(plain_password: str, hashed_password: str) -> bool:
    """Verify a plain-text password against a stored bcrypt hash."""
    return bcrypt.checkpw(
        plain_password.encode("utf-8"),
        hashed_password.encode("utf-8")
    )




#  DATABASE CREATION


def create_database():
    """Create all tables for the Smart Attendance System"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()


    # STUDENTS TABLE
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS students (
            student_id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            course TEXT NOT NULL,
            year INTEGER NOT NULL
        )
    """)


    # SUBJECTS TABLE
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS subjects (
            subject_id TEXT PRIMARY KEY,
            subject_name TEXT NOT NULL,
            course TEXT NOT NULL,
            year INTEGER NOT NULL,
            room TEXT NOT NULL,
            teacher_name TEXT NOT NULL
        )
    """)


    # ATTENDANCE TABLE
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id TEXT NOT NULL,
            subject_id TEXT NOT NULL,
            date TEXT NOT NULL,
            time TEXT NOT NULL,
            UNIQUE(student_id, subject_id, date),
            FOREIGN KEY(student_id) REFERENCES students(student_id),
            FOREIGN KEY(subject_id) REFERENCES subjects(subject_id)
        )
    """)


    # SESSIONS TABLE
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS sessions (
            session_id INTEGER PRIMARY KEY AUTOINCREMENT,
            subject_id TEXT NOT NULL,
            date TEXT NOT NULL,
            room TEXT NOT NULL,
            start_time TEXT,
            end_time TEXT,
            notes TEXT,
            UNIQUE(subject_id, date),
            FOREIGN KEY(subject_id) REFERENCES subjects(subject_id)
        )
    """)


    # ACTIVE_SESSIONS TABLE
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS active_sessions (
            session_id INTEGER PRIMARY KEY,
            opened_at TEXT NOT NULL,
            FOREIGN KEY(session_id) REFERENCES sessions(session_id)
        )
    """)


    # TEACHERS TABLE — passwords stored as bcrypt hashes
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS teachers (
            teacher_id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            email TEXT,
            phone TEXT,
            department TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)


    conn.commit()
    conn.close()
    print(f"✓ Database created: {DB_FILE}")
    print("  - students table")
    print("  - subjects table")
    print("  - attendance table")
    print("  - sessions table")
    print("  - active_sessions table")
    print("  - teachers table (bcrypt passwords)")




# ============ EXCEL IMPORT ============


def import_students_from_excel(filepath: str) -> dict:
    """
    Import students from an Excel file.


    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("✗ openpyxl not installed. Run: pip install openpyxl")
        return {"inserted": 0, "skipped": 0, "errors": ["openpyxl not installed"]}


    wb = load_workbook(filepath)
    ws = wb.active


    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()


    inserted = 0
    skipped = 0
    errors = []


    # Skip the header row
    rows = list(ws.iter_rows(min_row=2, values_only=True))


    for i, row in enumerate(rows, start=2):
        # Skip completely empty rows
        if not any(row):
            continue


        # Validate column count
        if len(row) < 4:
            errors.append(f"Row {i}: not enough columns (expected 4, got {len(row)})")
            continue


        student_id, name, course, year = row[0], row[1], row[2], row[3]


        # Validate required fields
        if not student_id or not name or not course or not year:
            errors.append(f"Row {i}: missing required field — {row}")
            continue


        try:
            year = int(year)
        except (ValueError, TypeError):
            errors.append(f"Row {i}: 'year' must be a number, got '{year}'")
            continue


        try:
            cursor.execute(
                "INSERT OR IGNORE INTO students (student_id, name, course, year) VALUES (?, ?, ?, ?)",
                (str(student_id).strip(), str(name).strip(), str(course).strip(), year)
            )
            if cursor.rowcount > 0:
                inserted += 1
            else:
                skipped += 1  # Student ID already exists
        except sqlite3.Error as e:
            errors.append(f"Row {i}: database error — {e}")


    conn.commit()
    conn.close()


    print(f"✓ Students import complete — inserted: {inserted}, skipped: {skipped}, errors: {len(errors)}")
    for err in errors:
        print(f"  ⚠ {err}")


    return {"inserted": inserted, "skipped": skipped, "errors": errors}




def import_teachers_from_excel(filepath: str) -> dict:
    """
    Import teachers from an Excel file.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("✗ openpyxl not installed. Run: pip install openpyxl")
        return {"inserted": 0, "skipped": 0, "errors": ["openpyxl not installed"]}


    wb = load_workbook(filepath)
    ws = wb.active


    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()


    inserted = 0
    skipped = 0
    errors = []


    rows = list(ws.iter_rows(min_row=2, values_only=True))


    for i, row in enumerate(rows, start=2):
        if not any(row):
            continue


        if len(row) < 6:
            errors.append(f"Row {i}: not enough columns (expected 6, got {len(row)})")
            continue


        name, email, phone, department, username, password = (
            row[0], row[1], row[2], row[3], row[4], row[5]
        )


        if not name or not username or not password:
            errors.append(f"Row {i}: name, username, and password are required")
            continue


        # Hash the plain-text password from the sheet
        hashed = hash_password(str(password))


        try:
            cursor.execute(
                """INSERT OR IGNORE INTO teachers
                   (name, username, password, email, phone, department)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (
                    str(name).strip(),
                    str(username).strip(),
                    hashed,
                    str(email).strip() if email else None,
                    str(phone).strip() if phone else None,
                    str(department).strip() if department else None,
                )
            )
            if cursor.rowcount > 0:
                inserted += 1
            else:
                skipped += 1  # Username already exists
        except sqlite3.Error as e:
            errors.append(f"Row {i}: database error — {e}")


    conn.commit()
    conn.close()


    print(f"✓ Teachers import complete — inserted: {inserted}, skipped: {skipped}, errors: {len(errors)}")
    for err in errors:
        print(f"  ⚠ {err}")


    return {"inserted": inserted, "skipped": skipped, "errors": errors}




# ============ STUDENT QUERY FUNCTIONS ============


def get_students_for_course(course, year):
    """Get all students in a course/year combination"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute(
        "SELECT student_id, name FROM students WHERE course = ? AND year = ?",
        (course, year)
    )
    result = cursor.fetchall()
    conn.close()
    return result




def get_subjects_for_student(student_id):
    """Get all subjects for a student's course/year"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT course, year FROM students WHERE student_id = ?", (student_id,))
    student = cursor.fetchone()
    if not student:
        conn.close()
        return []
    course, year = student
    cursor.execute(
        "SELECT subject_id, subject_name, room, teacher_name FROM subjects WHERE course = ? AND year = ?",
        (course, year)
    )
    result = cursor.fetchall()
    conn.close()
    return result




def get_all_subjects():
    """Get all subjects (for teacher view)"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT subject_id, subject_name, room, teacher_name FROM subjects")
    result = cursor.fetchall()
    conn.close()
    return result




def get_attendance_for_student_subject(student_id, subject_id):
    """Get all attendance records for a student in a specific subject"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute(
        "SELECT date, time FROM attendance WHERE student_id = ? AND subject_id = ? ORDER BY date DESC",
        (student_id, subject_id)
    )
    result = cursor.fetchall()
    conn.close()
    return result




def get_attendance_count_for_student_subject(student_id, subject_id):
    """Get count of days attended for a student in a subject"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute(
        "SELECT COUNT(DISTINCT date) FROM attendance WHERE student_id = ? AND subject_id = ?",
        (student_id, subject_id)
    )
    result = cursor.fetchone()[0]
    conn.close()
    return result




def get_total_class_days(subject_id):
    """Get number of unique class days for a subject"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute(
        "SELECT COUNT(DISTINCT date) FROM attendance WHERE subject_id = ?",
        (subject_id,)
    )
    result = cursor.fetchone()[0]
    conn.close()
    return result if result > 0 else 1




def get_todays_attendance(subject_id):
    """Get all students who have scanned in today for a subject"""
    today = str(date_class.today())
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute(
        """SELECT DISTINCT s.student_id, s.name, a.time
           FROM attendance a
           JOIN students s ON a.student_id = s.student_id
           WHERE a.subject_id = ? AND a.date = ?
           ORDER BY a.time ASC""",
        (subject_id, today)
    )
    result = cursor.fetchall()
    conn.close()
    return result




def get_full_attendance_report(subject_id):
    """Get full attendance summary for all students in a subject"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute(
        "SELECT DISTINCT course, year FROM subjects WHERE subject_id = ?",
        (subject_id,)
    )
    course_year = cursor.fetchone()
    if not course_year:
        conn.close()
        return []
    course, year = course_year
    cursor.execute(
        """SELECT s.student_id, s.name, COUNT(DISTINCT a.date) as days_present
           FROM students s
           LEFT JOIN attendance a ON s.student_id = a.student_id AND a.subject_id = ?
           WHERE s.course = ? AND s.year = ?
           GROUP BY s.student_id
           ORDER BY s.name ASC""",
        (subject_id, course, year)
    )
    result = cursor.fetchall()
    conn.close()
    return result




def student_exists(student_id):
    """Check if a student exists in the database"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT 1 FROM students WHERE student_id = ?", (student_id,))
    result = cursor.fetchone() is not None
    conn.close()
    return result




def get_student_name(student_id):
    """Get student name by ID"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM students WHERE student_id = ?", (student_id,))
    result = cursor.fetchone()
    conn.close()
    return result[0] if result else None




# ============ TEACHER AUTH FUNCTIONS ============


def verify_teacher_login(username, password):
    """
    Verify teacher login credentials using bcrypt.
    Returns teacher dict on success, None on failure.
    """
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute(
        "SELECT teacher_id, name, department, email, password FROM teachers WHERE username = ?",
        (username,)
    )
    result = cursor.fetchone()
    conn.close()


    if result and check_password(password, result[4]):
        return {
            "teacher_id": result[0],
            "name": result[1],
            "department": result[2],
            "email": result[3],
        }
    return None




def get_teacher_by_id(teacher_id):
    """Get teacher details by ID"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute(
        "SELECT teacher_id, name, username, email, phone, department FROM teachers WHERE teacher_id = ?",
        (teacher_id,)
    )
    result = cursor.fetchone()
    conn.close()
    if result:
        return {
            "teacher_id": result[0],
            "name": result[1],
            "username": result[2],
            "email": result[3],
            "phone": result[4],
            "department": result[5],
        }
    return None




def get_all_teachers():
    """Get all teachers"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT teacher_id, name, username, email, phone, department FROM teachers")
    result = cursor.fetchall()
    conn.close()
    return result




# ============ SESSION MANAGEMENT FUNCTIONS ============


def create_or_update_session(subject_id, date, room, start_time=None, end_time=None, notes=""):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    try:
        cursor.execute(
            """INSERT INTO sessions (subject_id, date, room, start_time, end_time, notes)
               VALUES (?, ?, ?, ?, ?, ?)
               ON CONFLICT(subject_id, date) DO UPDATE SET
                   room=excluded.room, start_time=excluded.start_time,
                   end_time=excluded.end_time, notes=excluded.notes""",
            (subject_id, date, room, start_time, end_time, notes)
        )
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"Error creating session: {e}")
        conn.close()
        return False




def get_session_for_date(subject_id, date):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute(
        """SELECT session_id, subject_id, date, room, start_time, end_time, notes
           FROM sessions WHERE subject_id = ? AND date = ?""",
        (subject_id, date)
    )
    result = cursor.fetchone()
    conn.close()
    return result




def get_room_for_subject_date(subject_id, date):
    session = get_session_for_date(subject_id, date)
    if session:
        return session[3]
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT room FROM subjects WHERE subject_id = ?", (subject_id,))
    result = cursor.fetchone()
    conn.close()
    return result[0] if result else None




def open_session(subject_id, date_str=None):
    if date_str is None:
        date_str = str(date_class.today())
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    session = get_session_for_date(subject_id, date_str)
    if not session:
        default_room = conn.execute(
            "SELECT room FROM subjects WHERE subject_id = ?", (subject_id,)
        ).fetchone()
        if not default_room:
            conn.close()
            return False
        cursor.execute(
            "INSERT INTO sessions (subject_id, date, room) VALUES (?, ?, ?)",
            (subject_id, date_str, default_room[0])
        )
        conn.commit()
        session = get_session_for_date(subject_id, date_str)
    if not session:
        conn.close()
        return False
    session_id = session[0]
    now = datetime.now().isoformat()
    try:
        cursor.execute(
            "INSERT INTO active_sessions (session_id, opened_at) VALUES (?, ?)",
            (session_id, now)
        )
        conn.commit()
        conn.close()
        return True
    except sqlite3.IntegrityError:
        conn.close()
        return False




def close_session(subject_id, date_str=None):
    if date_str is None:
        date_str = str(date_class.today())
    session = get_session_for_date(subject_id, date_str)
    if not session:
        return False
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM active_sessions WHERE session_id = ?", (session[0],))
    conn.commit()
    conn.close()
    return True




def is_session_open(subject_id, date_str=None):
    if date_str is None:
        date_str = str(date_class.today())
    session = get_session_for_date(subject_id, date_str)
    if not session:
        return False
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT 1 FROM active_sessions WHERE session_id = ?", (session[0],))
    result = cursor.fetchone() is not None
    conn.close()
    return result




def get_all_sessions_for_subject(subject_id, limit=30):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute(
        """SELECT session_id, subject_id, date, room, start_time, end_time, notes
           FROM sessions WHERE subject_id = ?
           ORDER BY date DESC LIMIT ?""",
        (subject_id, limit)
    )
    result = cursor.fetchall()
    conn.close()
    return result




if __name__ == "__main__":
    print("Initializing Smart Attendance System database...")
    create_database()
    print("Database ready.")
    print()
    print("To import data from Excel, run:")
    print("  from database import import_students_from_excel, import_teachers_from_excel")
    print("  import_students_from_excel('students.xlsx')")
    print("  import_teachers_from_excel('teachers.xlsx')")
